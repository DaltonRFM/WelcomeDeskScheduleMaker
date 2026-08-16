"""
Step 10: generates a colored .xlsx schedule matching the original
hand-built layout -- Day > Station column groups across the top, one
row per 15-min time slot down the side, each person's shift merged
into a colored block with their name.

No Google Sheets API / OAuth needed -- this is a plain local file you
can open directly in Excel/Numbers, or drag into Google Drive and open
with Sheets (File > Import > Upload in Sheets also works).

Usage (standalone):
    python -m src.xlsx_writer
(reads the already-generated data/generated_schedule.csv and renders it
-- run `python -m src.main` first to produce that CSV)
"""

from datetime import datetime, time, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models import Day, Shift, Station

SLOT_MINUTES = 15

# A palette of distinct, readable colors -- cycles if there are more
# people than colors (unlikely at ~20-25 people, but safe either way).
COLOR_PALETTE = [
    "1F4E78", "C00000", "ED7D31", "70AD47", "7030A0",
    "2E75B6", "BF9000", "548235", "A9D18E", "8497B0",
    "FF6699", "375623", "833C00", "203864", "9E480E",
    "D6B656", "674EA7", "A64D79", "45818E", "B45F06",
    "CC0000", "38761D", "0B5394", "351C75", "4C1130",
]


def _build_row_axis(day_operating_hours: dict) -> list:
    """One row per 15-min slot, spanning the EARLIEST start to LATEST
    end across all configured days (so a day with a later start, like
    Friday, just has blank rows at the top rather than its own axis)."""
    earliest_start = min(start for start, end in day_operating_hours.values())
    latest_end = max(end for start, end in day_operating_hours.values())

    rows = []
    current = datetime(2000, 1, 1, earliest_start.hour, earliest_start.minute)
    end_dt = datetime(2000, 1, 1, latest_end.hour, latest_end.minute)
    while current < end_dt:
        rows.append(current.time())
        current += timedelta(minutes=SLOT_MINUTES)
    return rows


def _assign_columns(shifts_for_station: list, capacity: int) -> list:
    """
    Greedy interval-packing: assigns each shift to a sub-column index
    (0..capacity-1) such that no two shifts sharing a column overlap in
    time. Since station capacity is a hard constraint the solver already
    respected, this should always succeed within `capacity` columns.
    Returns a list of (shift, column_index) tuples, in the same order
    as the input (Shift isn't hashable, so this avoids using it as a
    dict key).
    """
    sorted_shifts = sorted(shifts_for_station, key=lambda s: s.start)
    column_end_times = [None] * capacity  # last end time booked in each column
    assignment = []

    for shift in sorted_shifts:
        placed = False
        for col in range(capacity):
            if column_end_times[col] is None or column_end_times[col] <= shift.start:
                column_end_times[col] = shift.end
                assignment.append((shift, col))
                placed = True
                break
        if not placed:
            # Shouldn't happen if capacity was respected upstream, but
            # don't silently drop data -- extend with an overflow column.
            column_end_times.append(shift.end)
            assignment.append((shift, len(column_end_times) - 1))

    return assignment


def write_schedule_xlsx(
    shifts: list,
    day_operating_hours: dict,
    station_capacity: dict,
    filepath: str,
    station_order: list = None,
    rotation_stations: list = None,
) -> None:
    station_order = station_order or list(station_capacity.keys())
    days_in_order = [d for d in day_operating_hours.keys()]
    rotation_stations = rotation_stations or []

    row_axis = _build_row_axis(day_operating_hours)
    row_index = {t: i for i, t in enumerate(row_axis)}
    DATA_START_ROW = 3  # rows 1-2 are headers, time labels start row 3

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Assign each person a stable color
    people = sorted({s.person for s in shifts}, key=lambda p: p.name)
    person_color = {
        p: COLOR_PALETTE[i % len(COLOR_PALETTE)] for i, p in enumerate(people)
    }

    # Time label column
    ws.cell(row=1, column=1, value="")
    ws.cell(row=2, column=1, value="")
    for i, t in enumerate(row_axis):
        ws.cell(row=DATA_START_ROW + i, column=1, value=t.strftime("%I:%M %p").lstrip("0"))

    # Build column layout: for each day, for each station, capacity-many columns
    col = 2
    day_station_columns = {}  # (day, station) -> list of column indices
    for day in days_in_order:
        day_start_col = col
        for station in station_order:
            capacity = station_capacity[station]
            station_start_col = col
            day_station_columns[(day, station)] = list(range(col, col + capacity))
            col += capacity

            # Station header (row 2)
            if capacity > 1:
                ws.merge_cells(
                    start_row=2, start_column=station_start_col,
                    end_row=2, end_column=station_start_col + capacity - 1,
                )
            ws.cell(row=2, column=station_start_col, value=station.value)
            ws.cell(row=2, column=station_start_col).font = Font(bold=True)
            ws.cell(row=2, column=station_start_col).alignment = Alignment(horizontal="center")

        day_end_col = col - 1
        if day_end_col > day_start_col:
            ws.merge_cells(start_row=1, start_column=day_start_col, end_row=1, end_column=day_end_col)
        ws.cell(row=1, column=day_start_col, value=day.value)
        ws.cell(row=1, column=day_start_col).font = Font(bold=True, size=13)
        ws.cell(row=1, column=day_start_col).alignment = Alignment(horizontal="center")

    # Place each shift
    shifts_by_day_station = {}
    for s in shifts:
        shifts_by_day_station.setdefault((s.day, s.station), []).append(s)

    for (day, station), day_station_shifts in shifts_by_day_station.items():
        if (day, station) not in day_station_columns:
            continue  # day/station not in this run's config, skip defensively
        capacity = station_capacity[station]
        col_assignment = _assign_columns(day_station_shifts, capacity)
        available_columns = day_station_columns[(day, station)]

        for shift, col_offset in col_assignment:
            target_col = available_columns[min(col_offset, len(available_columns) - 1)]
            start_row = DATA_START_ROW + row_index[shift.start]
            # end_row: last row covered is the slot BEFORE shift.end
            end_row = DATA_START_ROW + row_index[shift.start] + round(
                (datetime(2000, 1, 1, shift.end.hour, shift.end.minute)
                 - datetime(2000, 1, 1, shift.start.hour, shift.start.minute)).seconds / 60 / SLOT_MINUTES
            ) - 1

            if end_row >= start_row:
                if end_row > start_row:
                    ws.merge_cells(
                        start_row=start_row, start_column=target_col,
                        end_row=end_row, end_column=target_col,
                    )
                cell = ws.cell(row=start_row, column=target_col)
                cell.value = f"{shift.person.name}\n({shift.start.strftime('%I:%M').lstrip('0')}-{shift.end.strftime('%I:%M %p').lstrip('0')})"
                cell.fill = PatternFill(
                    start_color=person_color[shift.person],
                    end_color=person_color[shift.person],
                    fill_type="solid",
                )
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Summary section, below the main grid ---
    # Matches the original hand-built sheet: one row per person, hours
    # worked per day + weekly total, then checkmarks for opening,
    # closing, and each rotation station worked at least once.
    summary_start_row = DATA_START_ROW + len(row_axis) + 2

    CHECK = "\u2705"  # green checkmark, matches the original sheet's style
    CROSS = "x"

    summary_headers = ["Name"] + [d.value for d in days_in_order] + ["Total", "Opening", "Closing"] + [s.value for s in rotation_stations]
    for i, header in enumerate(summary_headers):
        cell = ws.cell(row=summary_start_row, column=1 + i, value=header)
        cell.font = Font(bold=True)

    for row_offset, person in enumerate(people):
        row = summary_start_row + 1 + row_offset
        person_shifts = [s for s in shifts if s.person == person]

        # Name (colored to match their blocks above, like the original)
        name_cell = ws.cell(row=row, column=1, value=person.name)
        name_cell.fill = PatternFill(
            start_color=person_color[person], end_color=person_color[person], fill_type="solid"
        )
        name_cell.font = Font(color="FFFFFF", bold=True)

        total_hours = 0.0
        for day_offset, day in enumerate(days_in_order):
            day_hours = sum(s.duration_hours() for s in person_shifts if s.day == day)
            total_hours += day_hours
            if day_hours > 0:
                ws.cell(row=row, column=2 + day_offset, value=day_hours)

        total_col = 2 + len(days_in_order)
        ws.cell(row=row, column=total_col, value=total_hours).font = Font(bold=True)

        # Opening / closing: did they work the FIRST/LAST slot of ANY day
        opened = any(
            s.day in day_operating_hours and s.start == day_operating_hours[s.day][0]
            for s in person_shifts
        )
        closed = any(
            s.day in day_operating_hours and s.end == day_operating_hours[s.day][1]
            for s in person_shifts
        )
        ws.cell(row=row, column=total_col + 1, value=CHECK if opened else CROSS)
        ws.cell(row=row, column=total_col + 2, value=CHECK if closed else CROSS)

        # One column per rotation station: did they work it at least once
        for station_offset, station in enumerate(rotation_stations):
            worked = any(s.station == station for s in person_shifts)
            ws.cell(
                row=row, column=total_col + 3 + station_offset,
                value=CHECK if worked else CROSS,
            )

    # Reasonable column widths
    ws.column_dimensions["A"].width = 10
    for c in range(2, col):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(filepath)


if __name__ == "__main__":
    from src.main import DAY_OPERATING_HOURS, STATION_CAPACITY, CSV_FILES, ROTATION_STATIONS
    from src.output_writer import read_schedule_csv

    schedule_csv_path = "data/generated_schedule.csv"
    active_day_hours = {
        day: hours for day, hours in DAY_OPERATING_HOURS.items() if day in CSV_FILES
    }

    try:
        shifts = read_schedule_csv(schedule_csv_path)
    except FileNotFoundError:
        print(f"{schedule_csv_path} not found -- run `python -m src.main` first to generate it.")
    else:
        output_path = "data/generated_schedule.xlsx"
        write_schedule_xlsx(
            shifts, active_day_hours, STATION_CAPACITY, output_path,
            rotation_stations=ROTATION_STATIONS,
        )
        print(f"Colored schedule written to {output_path}")